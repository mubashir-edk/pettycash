from django.shortcuts import render, redirect
from django.urls import reverse
from .models import *
from .forms import *
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
from PIL import Image
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Image, PageBreak, Frame, PageTemplate
import base64
from openpyxl import Workbook
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment



# First Page-----------------------------------------------------------------------------------------------------------------------------------------
@login_required
def chooseCashBook(request):
    return render(request, 'main.html')

# FetchClosingBalance--------------------------------------------------------------------------------------------------------------------------------
@login_required
def fetchOpeningBalance(request):
        
    try:
        op_balance = OpeningBalance.objects.filter(opening_balance_of="PettyCash").first()
        
        if op_balance:
            
            op_balance_amount = op_balance.opening_balance
            
            pettycashes = PettyCash.objects.all()
            
            pettycash_closing_balance = 0
            
            for pettycash in pettycashes:
                
                if pettycash.cash_flow == "Cash In":
                    pettycash_closing_balance = pettycash_closing_balance + pettycash.amount
                else:
                    pettycash_closing_balance = pettycash_closing_balance - pettycash.amount
                
            closing_balance = float(op_balance_amount) + pettycash_closing_balance
            
            closing_balance = Decimal(closing_balance).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            
            data = {'closing_balance': {
            'balance': closing_balance,
            }}
            
        else:
            
            data = {'closing_balance': {
            'balance': '0.00',
            }}
        
        return JsonResponse(data)
    except OpeningBalance.DoesNotExist:
        return JsonResponse({'error': 'Balance not found'}, status=404)

@login_required
def fetchFederalOpeningBalance(request):
        
    try:
        op_balance = OpeningBalance.objects.filter(opening_balance_of="Federal Bank").first()
        
        if op_balance:
            
            op_balance_amount = op_balance.opening_balance
            
            federal_pettycashes = FederalBank.objects.all()
            
            pettycash_closing_balance = 0
            
            for federal_pettycash in federal_pettycashes:
                
                if federal_pettycash.cash_flow == "Cash In":
                    pettycash_closing_balance = pettycash_closing_balance + federal_pettycash.amount
                else:
                    pettycash_closing_balance = pettycash_closing_balance - federal_pettycash.amount
                
            closing_balance = float(op_balance_amount) + pettycash_closing_balance
            
            closing_balance = Decimal(closing_balance).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            
            data = {'closing_balance': {
            'balance': closing_balance,
            }}
            
        else:
            
            data = {'closing_balance': {
            'balance': '0.00',
            }}
        
        return JsonResponse(data)
    except OpeningBalance.DoesNotExist:
        return JsonResponse({'error': 'Balance not found'}, status=404)

@login_required
def fetchSBIOpeningBalance(request):
        
    try:
        op_balance = OpeningBalance.objects.filter(opening_balance_of="SBI Bank").first()
        
        if op_balance:
            
            op_balance_amount = op_balance.opening_balance
            
            sbi_pettycashes = SBIBank.objects.all()
            
            pettycash_closing_balance = 0
            
            for sbi_pettycash in sbi_pettycashes:
                
                if sbi_pettycash.cash_flow == "Cash In":
                    pettycash_closing_balance = pettycash_closing_balance + sbi_pettycash.amount
                else:
                    pettycash_closing_balance = pettycash_closing_balance - sbi_pettycash.amount
                
            closing_balance = float(op_balance_amount) + pettycash_closing_balance
            
            closing_balance = Decimal(closing_balance).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            
            data = {'closing_balance': {
            'balance': closing_balance,
            }}
            
        else:
            
            data = {'closing_balance': {
            'balance': '0.00',
            }}
        
        return JsonResponse(data)
    except OpeningBalance.DoesNotExist:
        return JsonResponse({'error': 'Balance not found'}, status=404)
    

# Category-------------------------------------------------------------------------------------------------------------------------------------------
@login_required
@staff_member_required
def viewCategory(request):
    
    categories = Category.objects.all()
    category_exists = categories.exists()
    category_form = CategoryForm()
    context = {'categories': categories, 'category_exists': category_exists, 'category_form': category_form}
    return render(request, 'category.html', context)

@login_required
@staff_member_required
def createCategory(request):
    
    print("in url")
    
    if request.method == 'POST':
        
        print("in post")
        
        category_form = CategoryForm(request.POST)
        if category_form.is_valid():
            print("in valid")
            
            category_form.save()
            return redirect('pettycash_app:category')

@login_required
@staff_member_required
def updateCategory(request, id):
    
    category = get_object_or_404(Category, pk=id)
    category_form = CategoryForm(instance=category)
    if request.method == 'POST':
        
        category_form = CategoryForm(request.POST, instance=category)
        if category_form.is_valid():
            
            category_form.save()
            return redirect('pettycash_app:category')
            
@login_required
@staff_member_required
def deleteCategory(request, id):
    
    category = get_object_or_404(Category, pk=id)
    category.delete()
    return redirect('pettycash_app:category')


# PettyCash-------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def viewPettyCash(request):
    
    pettycashes = PettyCash.objects.all().order_by('-voucher_code')
    pettycash_exists = pettycashes.exists()
    context = {'pettycashes': pettycashes, 'pettycash_exists': pettycash_exists}
    return render(request, 'pettycash/view_pettycash.html', context)

@login_required
def createPettyCash(request):
    
    pettycash_form = PettyCashForm()
    if request.method == 'POST':
        
        pettycash_form = PettyCashForm(request.POST)
        
        # Generating Voucher Code
        voucher_codes_only = VoucherCodeStore.objects.all().values_list('voucher_code', flat=True)
        
        voucher_code_loop = True
        voucher_code_number = 1001
        
        while voucher_code_loop:
            
            generated_voucher_code = "VC_" + str(voucher_code_number)

            if generated_voucher_code in voucher_codes_only:
                voucher_code_number += 1
            else:
                break
                
        if pettycash_form.is_valid():
            
            pettycash = pettycash_form.save(commit=False)
            pettycash.voucher_code = generated_voucher_code
            VoucherCodeStore.objects.create(voucher_code=generated_voucher_code)
            pettycash.save()
            return redirect('pettycash_app:view_pettycash')
            
    context = {'pettycash_form': pettycash_form}
    return render(request, 'pettycash/pettycash.html', context)

@login_required
@staff_member_required
def updatePettyCash(request, id):
    
    pettycash = get_object_or_404(PettyCash, pk=id)
    voucher_code_stored =  pettycash.voucher_code
    pettycash_form = PettyCashForm(instance=pettycash)
    if request.method == 'POST':
        
        pettycash_form = PettyCashForm(request.POST, instance=pettycash)
        if pettycash_form.is_valid():
            pettycash = pettycash_form.save(commit=False)
            pettycash.voucher_code = voucher_code_stored
            pettycash.save()
            return redirect(reverse('pettycash_app:each_pettycash', kwargs={'id': pettycash.id}))
            
    context = {'pettycash_form': pettycash_form, 'pettycash': pettycash}
    return render(request, 'pettycash/pettycash.html', context)

@login_required
def eachPettyCash(request, id):
    
    pettycash = get_object_or_404(PettyCash, pk=id)
    context = {'pettycash': pettycash}
    return render(request, 'pettycash/each_pettycash.html', context)

@login_required
@staff_member_required
def pettyCashSignature(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        image_data = request.POST.get('image_data')

        print('pettycash', image_data)
    
        if image_data:
            # Decode base64 image data
            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            image_data = ContentFile(base64.b64decode(imgstr), name='temp.' + ext)

            # Assuming 'pettycash' is your model instance
            pettycash = PettyCash.objects.get(pk=request.POST.get('pettycash_id'))  # Adjust this according to your model

            # Save image data to model field
            pettycash.image.save('signature.' + ext, image_data, save=True)

            return JsonResponse({'success': True})  # Send JSON response indicating success
        else:
            return JsonResponse({'success': False, 'error': 'No image data received'})  # Handle no image data received
    else:
        return HttpResponseBadRequest('Invalid request') 

@login_required
@staff_member_required
def deletePettyCash(request, id):
    
    pettycash = get_object_or_404(PettyCash, pk=id)
    pettycash.delete()
    return redirect('pettycash_app:view_pettycash')


# FederalBankPettyCash-------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def viewFederalPettyCash(request):
    
    federal_pettycashes = FederalBank.objects.all().order_by('-voucher_code')
    federal_pettycash_exists = federal_pettycashes.exists()
    context = {'federal_pettycashes': federal_pettycashes, 'federal_pettycash_exists': federal_pettycash_exists}
    return render(request, 'federal_bank/view_federal_pettycash.html', context)

@login_required
def createFederalPettyCash(request):
    
    federal_pettycash_form = FederalBankPettyCashForm()
    if request.method == 'POST':
        
        federal_pettycash_form = FederalBankPettyCashForm(request.POST)
        
        # Generating Voucher Code
        voucher_codes_only = VoucherCodeStore.objects.all().values_list('voucher_code', flat=True)
        
        voucher_code_loop = True
        voucher_code_number = 1001
        
        while voucher_code_loop:
            
            generated_voucher_code = "VC_" + str(voucher_code_number)

            if generated_voucher_code in voucher_codes_only:
                voucher_code_number += 1
            else:
                break
                
        if federal_pettycash_form.is_valid():
            
            federal_pettycash = federal_pettycash_form.save(commit=False)
            federal_pettycash.voucher_code = generated_voucher_code
            VoucherCodeStore.objects.create(voucher_code=generated_voucher_code)
            federal_pettycash.save()
            return redirect('pettycash_app:view_federal_pettycash')
            
    context = {'federal_pettycash_form': federal_pettycash_form}
    return render(request, 'federal_bank/federal_pettycash.html', context)

@login_required
@staff_member_required
def updateFederalPettyCash(request, id):
    
    federal_pettycash = get_object_or_404(FederalBank, pk=id)
    voucher_code_stored =  federal_pettycash.voucher_code
    federal_pettycash_form = PettyCashForm(instance=federal_pettycash)
    if request.method == 'POST':
        
        federal_pettycash_form = PettyCashForm(request.POST, instance=federal_pettycash)
        if federal_pettycash_form.is_valid():
            federal_pettycash_save = federal_pettycash_form.save(commit=False)
            federal_pettycash_save.voucher_code = voucher_code_stored
            federal_pettycash_save.save()
            return redirect(reverse('pettycash_app:each_federal_pettycash', kwargs={'id': federal_pettycash.id}))
            
    context = {'federal_pettycash_form': federal_pettycash_form, 'federal_pettycash': federal_pettycash}
    return render(request, 'federal_bank/federal_pettycash.html', context)

@login_required
def eachFederalPettyCash(request, id):
    
    federal_pettycash = get_object_or_404(FederalBank, pk=id)
    context = {'federal_pettycash': federal_pettycash}
    return render(request, 'federal_bank/each_federal_pettycash.html', context)

@login_required
@staff_member_required
def federalBankSignature(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        image_data = request.POST.get('image_data')

        print('pettycash', image_data)
    
        if image_data:
            # Decode base64 image data
            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            image_data = ContentFile(base64.b64decode(imgstr), name='temp.' + ext)

            # Assuming 'pettycash' is your model instance
            federal_pettycash = FederalBank.objects.get(pk=request.POST.get('federal_pettycash_id'))  # Adjust this according to your model

            # Save image data to model field
            federal_pettycash.image.save('signature.' + ext, image_data, save=True)

            return JsonResponse({'success': True})  # Send JSON response indicating success
        else:
            return JsonResponse({'success': False, 'error': 'No image data received'})  # Handle no image data received
    else:
        return HttpResponseBadRequest('Invalid request') 

@login_required
@staff_member_required
def deleteFederalPettyCash(request, id):
    
    federal_pettycash = get_object_or_404(FederalBank, pk=id)
    federal_pettycash.delete()
    return redirect('pettycash_app:view_federal_pettycash')


# SBIBankPettyCash-------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def viewSBIPettyCash(request):
    
    sbi_pettycashes = SBIBank.objects.all().order_by('-voucher_code')
    sbi_pettycash_exists = sbi_pettycashes.exists()
    context = {'sbi_pettycashes': sbi_pettycashes, 'sbi_pettycash_exists': sbi_pettycash_exists}
    return render(request, 'sbi_bank/view_sbi_pettycash.html', context)

@login_required
def createSBIPettyCash(request):
    
    sbi_pettycash_form = SBIBankPettyCashForm()
    if request.method == 'POST':
        
        sbi_pettycash_form = SBIBankPettyCashForm(request.POST)
        
        # Generating Voucher Code
        voucher_codes_only = VoucherCodeStore.objects.all().values_list('voucher_code', flat=True)
        
        voucher_code_loop = True
        voucher_code_number = 1001
        
        while voucher_code_loop:
            
            generated_voucher_code = "VC_" + str(voucher_code_number)

            if generated_voucher_code in voucher_codes_only:
                voucher_code_number += 1
            else:
                break
                
        if sbi_pettycash_form.is_valid():
            
            sbi_pettycash = sbi_pettycash_form.save(commit=False)
            sbi_pettycash.voucher_code = generated_voucher_code
            VoucherCodeStore.objects.create(voucher_code=generated_voucher_code)
            sbi_pettycash.save()
            return redirect('pettycash_app:view_sbi_pettycash')
            
    context = {'sbi_pettycash_form': sbi_pettycash_form}
    return render(request, 'sbi_bank/sbi_pettycash.html', context)

@login_required
@staff_member_required
def updateSBIPettyCash(request, id):
    
    sbi_pettycash = get_object_or_404(SBIBank, pk=id)
    voucher_code_stored =  sbi_pettycash.voucher_code
    sbi_pettycash_form = SBIBankPettyCashForm(instance=sbi_pettycash)
    if request.method == 'POST':
        
        sbi_pettycash_form = SBIBankPettyCashForm(request.POST, instance=sbi_pettycash)
        if sbi_pettycash_form.is_valid():
            federal_pettycash_save = sbi_pettycash_form.save(commit=False)
            federal_pettycash_save.voucher_code = voucher_code_stored
            federal_pettycash_save.save()
            return redirect(reverse('pettycash_app:each_sbi_pettycash', kwargs={'id': sbi_pettycash.id}))
            
    context = {'sbi_pettycash_form': sbi_pettycash_form, 'sbi_pettycash': sbi_pettycash}
    return render(request, 'sbi_bank/sbi_pettycash.html', context)

@login_required
def eachSBIPettyCash(request, id):
    
    sbi_pettycash = get_object_or_404(SBIBank, pk=id)
    context = {'sbi_pettycash': sbi_pettycash}
    return render(request, 'sbi_bank/each_sbi_pettycash.html', context)

@login_required
@staff_member_required
def SbiBankSignature(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        image_data = request.POST.get('image_data')

        print('pettycash', image_data)
    
        if image_data:
            # Decode base64 image data
            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            image_data = ContentFile(base64.b64decode(imgstr), name='temp.' + ext)

            # Assuming 'pettycash' is your model instance
            sbi_pettycash = SBIBank.objects.get(pk=request.POST.get('sbi_bank_id'))  # Adjust this according to your model

            # Save image data to model field
            sbi_pettycash.image.save('signature.' + ext, image_data, save=True)

            return JsonResponse({'success': True})  # Send JSON response indicating success
        else:
            return JsonResponse({'success': False, 'error': 'No image data received'})  # Handle no image data received
    else:
        return HttpResponseBadRequest('Invalid request')

@login_required
@staff_member_required
def deleteSBIPettyCash(request, id):
    
    sbi_pettycash = get_object_or_404(SBIBank, pk=id)
    sbi_pettycash.delete()
    return redirect('pettycash_app:view_sbi_pettycash')


# OpeningBalance-------------------------------------------------------------------------------------------------------------------------------------
@login_required
@staff_member_required
def viewOpeningBalance(request):
    
    open_balance_form = OpeningBalanceForm()
    open_balances = OpeningBalance.objects.all()
    open_balances_exists = open_balances.exists()
    context = {'open_balances': open_balances, 'open_balance_form': open_balance_form, 'open_balances_exists': open_balances_exists}
    return render(request, 'opening_balances.html', context)

@login_required
@staff_member_required
def createOpeningBalance(request):
    
    open_balance_form = OpeningBalanceForm()
    
    if request.method == 'POST':
        
        open_balance_form = OpeningBalanceForm(request.POST)
        
        if open_balance_form.is_valid():
            
            open_balance_form.save()
            
            return redirect('pettycash_app:view_opening_balance')

@login_required
@staff_member_required
def updateOpeningBalance(request, id):
    
    op = get_object_or_404(OpeningBalance, pk=id)
    
    open_balance_form = OpeningBalanceForm(instance=op)
    
    if request.method == 'POST':
        
        open_balance_form = OpeningBalanceForm(request.POST, instance=op)
        
        if open_balance_form.is_valid():
            open_balance_form.save()
            
            return redirect('pettycash_app:view_opening_balance')

@login_required
@staff_member_required
def deleteOpeningBalance(request, id):
    
    op = get_object_or_404(OpeningBalance, pk=id)
    op.delete()
    return redirect('pettycash_app:view_opening_balance')



# GenerateExcel--------------------------------------------------------------------------------------------------------------------------------------
@login_required
@staff_member_required
def pettycash_balance_sheet_excel(request):
    # Get the opening balance
    opening_balance = OpeningBalance.objects.filter(opening_balance_of='PettyCash').first()

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers
    ws.append(['PettyCash Balance Sheet', '', '', '', '', ''])
    ws.merge_cells('A{}:F{}'.format(ws.min_row, ws.min_row))
    
    for first_row in ws[ws.min_row]:
        first_row.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.append(['Date', 'Paid/Recieved', '(Cr.)', '(Dr.)', 'Balance', 'Description'])
    
    opening_date = opening_balance.opening_date.strftime('%d-%b-%Y')
    
    ws.append([opening_date ,'Opening Balance', '', '', opening_balance.opening_balance, ''])

    # Initialize balance
    balance = opening_balance.opening_balance if opening_balance else 0

    # Get all categories
    pettycashes = PettyCash.objects.all().order_by('created_date')

    # Iterate over each category
    for pettycash in pettycashes:
        # Calculate total cash-in and cash-out for the category
        cash_in_total = pettycash.amount if pettycash.cash_flow == 'Cash In' else 0
        cash_out_total = pettycash.amount if pettycash.cash_flow == 'Cash Out' else 0
        
        # Update balance
        balance += Decimal(cash_in_total) - Decimal(cash_out_total)
        
        balance.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        
        pettycash_date = pettycash.created_date.strftime('%d-%b-%Y')

        # Add data to the worksheet
        ws.append([pettycash_date, pettycash.to_or_from, cash_in_total, cash_out_total, balance, pettycash.description])
        
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        
    ws.append(['', '', '', '', balance, ''])

    # Save the workbook in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pettycash_balance_sheet.xlsx"'
    response.write(output.getvalue())

    return response

@login_required
@staff_member_required
def federal_balance_sheet_excel(request):
    # Get the opening balance
    opening_balance = OpeningBalance.objects.filter(opening_balance_of='Federal Bank').first()

     # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers
    ws.append(['Federal Bank Balance Sheet', '', '', '', '', ''])
    ws.merge_cells('A{}:F{}'.format(ws.min_row, ws.min_row))
    
    for first_row in ws[ws.min_row]:
        first_row.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.append(['Date', 'Paid/Recieved', '(Cr.)', '(Dr.)', 'Balance', 'Description'])
    
    opening_date = opening_balance.opening_date.strftime('%d-%b-%Y')
    ws.append([opening_date ,'Opening Balance', '', '', opening_balance.opening_balance, ''])

    # Initialize balance
    balance = opening_balance.opening_balance if opening_balance else 0

    # Get all categories
    federal_bank = FederalBank.objects.all().order_by('created_date')

    # Iterate over each category
    for federal in federal_bank:
        # Calculate total cash-in and cash-out for the category
        cash_in_total = federal.amount if federal.cash_flow == 'Cash In' else 0
        cash_out_total = federal.amount if federal.cash_flow == 'Cash Out' else 0
        
        # Update balance
        balance += Decimal(cash_in_total) - Decimal(cash_out_total)
        
        balance.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

        federal_date = federal.created_date.strftime('%d-%b-%Y')
        # Add data to the worksheet
        ws.append([federal_date, federal.to_or_from, cash_in_total, cash_out_total, balance, federal.description])
        
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        
    ws.append(['', '', '', '', balance, ''])

    # Save the workbook in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="federalbank_balance_sheet.xlsx"'
    response.write(output.getvalue())

    return response

@login_required
@staff_member_required
def sbi_balance_sheet_excel(request):
    # Get the opening balance
    opening_balance = OpeningBalance.objects.filter(opening_balance_of='SBI Bank').first()

     # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers
    ws.append(['SBI Bank Balance Sheet', '', '', '', '', ''])
    ws.merge_cells('A{}:F{}'.format(ws.min_row, ws.min_row))
    
    for first_row in ws[ws.min_row]:
        first_row.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.append(['Date', 'Paid/Recieved', '(Cr.)', '(Dr.)', 'Balance', 'Description'])
    
    opening_date = opening_balance.opening_date.strftime('%d-%b-%Y')
    ws.append([opening_date ,'Opening Balance', '', '', opening_balance.opening_balance, ''])

    # Initialize balance
    balance = opening_balance.opening_balance if opening_balance else 0

    # Get all categories
    sbi_bank = SBIBank.objects.all().order_by('created_date')

    # Iterate over each category
    for sbi in sbi_bank:
        # Calculate total cash-in and cash-out for the category
        cash_in_total = sbi.amount if sbi.cash_flow == 'Cash In' else 0
        cash_out_total = sbi.amount if sbi.cash_flow == 'Cash Out' else 0
        
        # Update balance
        balance += Decimal(cash_in_total) - Decimal(cash_out_total)
        
        balance.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

        sbi_date = sbi.created_date.strftime('%d-%b-%Y')
        # Add data to the worksheet
        ws.append([sbi_date, sbi.to_or_from, cash_in_total, cash_out_total, balance, sbi.description])
        
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        
    ws.append(['', '', '', '', balance, ''])

    # Save the workbook in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sbibank_balance_sheet.xlsx"'
    response.write(output.getvalue())

    return response


# Filtering view----------------------------------------------------------------------------------------------------------------------------------------
@login_required
def filter_federal_pettycash(request):
    # Fetching the filter parameters from the request
    from_date = request.GET.get('fromDate')
    to_date = request.GET.get('toDate')
    cash_type = request.GET.get('cashType')

    # Start with the full queryset
    federal_pettycashes = FederalBank.objects.all()

    # Apply filters if provided
    if from_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        federal_pettycashes = federal_pettycashes.filter(created_date__gte=from_date)
    if to_date:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        federal_pettycashes = federal_pettycashes.filter(created_date__lte=to_date)
    if cash_type != 'All':
        federal_pettycashes = federal_pettycashes.filter(cash_flow=cash_type)

    # Order the queryset
    federal_pettycashes = federal_pettycashes.order_by('-voucher_code')

    # Serialize queryset to JSON with category name
    data = []
    for cash in federal_pettycashes:
        serialized_data = {
            'id': cash.id,
            'voucher_code': cash.voucher_code,
            'created_date': cash.created_date,
            'category': cash.category.category,  # Access category name directly
            'cash_flow': cash.cash_flow,
            'amount': cash.amount,
            'description': cash.description
        }
        data.append(serialized_data)

    if data:
        data_exists = True
    else:
        data_exists = False

    request.session['filtering_complete'] = True
    
    result = {'data': data, 'data_exists': data_exists}

    return JsonResponse(result, safe=False)

@login_required
def filter_pettycash(request):
    # Fetching the filter parameters from the request
    from_date = request.GET.get('fromDate')
    to_date = request.GET.get('toDate')
    cash_type = request.GET.get('cashType')

    # Start with the full queryset
    pettycashes = PettyCash.objects.all()

    # Apply filters if provided
    if from_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        pettycashes = pettycashes.filter(created_date__gte=from_date)
    if to_date:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        pettycashes = pettycashes.filter(created_date__lte=to_date)
    if cash_type != 'All':
        pettycashes = pettycashes.filter(cash_flow=cash_type)

    # Order the queryset
    pettycashes = pettycashes.order_by('-voucher_code')

    # Serialize queryset to JSON with category name
    data = []
    for cash in pettycashes:
        serialized_data = {
            'id': cash.id,
            'voucher_code': cash.voucher_code,
            'created_date': cash.created_date,
            'category': cash.category.category,  # Access category name directly
            'cash_flow': cash.cash_flow,
            'amount': cash.amount,
            'description': cash.description
        }
        data.append(serialized_data)

    if data:
        data_exists = True
    else:
        data_exists = False

    request.session['filtering_complete'] = True
    
    result = {'data': data, 'data_exists': data_exists}

    return JsonResponse(result, safe=False)

@login_required
def filter_sbi_pettycash(request):
    # Fetching the filter parameters from the request
    from_date = request.GET.get('fromDate')
    to_date = request.GET.get('toDate')
    cash_type = request.GET.get('cashType')

    # Start with the full queryset
    sbi_pettycashes = SBIBank.objects.all()

    # Apply filters if provided
    if from_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        sbi_pettycashes = sbi_pettycashes.filter(created_date__gte=from_date)
    if to_date:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        sbi_pettycashes = sbi_pettycashes.filter(created_date__lte=to_date)
    if cash_type != 'All':
        sbi_pettycashes = sbi_pettycashes.filter(cash_flow=cash_type)

    # Order the queryset
    sbi_pettycashes = sbi_pettycashes.order_by('-voucher_code')

    # Serialize queryset to JSON with category name
    data = []
    for cash in sbi_pettycashes:
        serialized_data = {
            'id': cash.id,
            'voucher_code': cash.voucher_code,
            'created_date': cash.created_date,
            'category': cash.category.category,  # Access category name directly
            'cash_flow': cash.cash_flow,
            'amount': cash.amount,
            'description': cash.description
        }
        data.append(serialized_data)

    if data:
        data_exists = True
    else:
        data_exists = False

    request.session['filtering_complete'] = True
    
    result = {'data': data, 'data_exists': data_exists}

    return JsonResponse(result, safe=False)


# Generate PDF----------------------------------------------------------------------------------------------------------------------------------------
@login_required
def pettycash_all_pdf(request):
    # pettycashes = PettyCash.objects.all()
    
    pettycash_type = request.POST.get('allTypePettycash')
        
    if pettycash_type == "pettycash":
        pettycashes = PettyCash.objects.all().order_by('voucher_code')
    elif pettycash_type == "federal_bank":
        pettycashes = FederalBank.objects.all().order_by('voucher_code')
    elif pettycash_type == "sbi_bank":
        pettycashes = SBIBank.objects.all().order_by('voucher_code')
    else:
        return f"None"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=10, rightMargin=10, topMargin=10, bottomMargin=10, title="All Petty Cashes")
    elements = []
    table_count = 0

    for pettycash in pettycashes:
        data = [
            ["Voucher Code:", pettycash.voucher_code],
            ["Date:", pettycash.created_date],
            ["To/From:", pettycash.to_or_from],
            ["Type:", pettycash.cash_flow],
            ["Category:", pettycash.category],
            ["Amount:", pettycash.amount],
            ["Description:", pettycash.description],
            ["Signature:", Image(pettycash.image.path, width=90, height=50) if pettycash.image else "Not Signed"],  # Embed image if available
        ]

        # Create table
        table = Table(data, colWidths=[120, '*'])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Gray background for header row
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Align text to the left
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to the top
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Text color
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),  # Font
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Inner grid
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Border
        ]))

        elements.append(table)
        elements.append(Spacer(1, 12))  # Add spacing between tables
        table_count += 1
        
        if table_count % 4 == 0:
            elements.append(PageBreak())

    # If there are remaining elements, build the final page
    if elements:
        doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pettycashes.pdf"'
    response.write(pdf)
    return response

@login_required
def pettycash_custom_pdf(request):
    
    if request.method == 'POST':
        
        selected_voucher = request.POST.get('VoucherCodesSelect')
        
        pettycash_type = request.POST.get('pettycashType')
        
        if pettycash_type == "pettycash":
            pettycashes = PettyCash.objects.filter(voucher_code__gte=selected_voucher).order_by('voucher_code')
        elif pettycash_type == "federal_bank":
            pettycashes = FederalBank.objects.filter(voucher_code__gte=selected_voucher).order_by('voucher_code')
        elif pettycash_type == "sbi_bank":
            pettycashes = SBIBank.objects.filter(voucher_code__gte=selected_voucher).order_by('voucher_code')
        else:
            return f"None"
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=10, rightMargin=10, topMargin=10, bottomMargin=10, title="All Petty Cashes")
        elements = []
        table_count = 0

        for pettycash in pettycashes:
            data = [
                ["Voucher Code:", pettycash.voucher_code],
                ["Date:", pettycash.created_date],
                ["To/From:", pettycash.to_or_from],
                ["Type:", pettycash.cash_flow],
                ["Category:", pettycash.category],
                ["Amount:", pettycash.amount],
                ["Description:", pettycash.description],
                ["Signature:", Image(pettycash.image.path, width=90, height=50) if pettycash.image else "Not Signed"],  # Embed image if available
            ]

            # Create table
            table = Table(data, colWidths=[120, '*'])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Gray background for header row
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Align text to the left
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to the top
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Text color
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),  # Font
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Inner grid
                ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Border
            ]))

            elements.append(table)
            elements.append(Spacer(1, 12))  # Add spacing between tables
            table_count += 1
            
            if table_count % 4 == 0:
                elements.append(PageBreak())

        # If there are remaining elements, build the final page
        if elements:
            doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="selected_pettycashes.pdf"'
        response.write(pdf)
        return response